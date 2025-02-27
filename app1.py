from flask import Flask, render_template, request, redirect, url_for
import os
import openpyxl

app = Flask(__name__)

# Path to the Excel file
FILE_PATH = os.path.join(os.getcwd(), "log_history.xlsx")

# Create an Excel file if it doesn't exist
if not os.path.exists(FILE_PATH):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "History"
    ws.append(["Date", "Episode/Chapter", "Title", "Language", "Duration", "Watched/Read At", "Comment/Review"])
    wb.save(FILE_PATH)

def format_episodes(episode):
    # Split by commas to handle formats like "2-4,6,9-10"
    episode_ranges = episode.split(',')
    formatted_episodes = []

    for ep in episode_ranges:
        # Check if it's a range (e.g., "2-4")
        if '-' in ep:
            start, end = map(int, ep.split('-'))
            # Generate the range (e.g., "2-4")
            formatted_episodes.append(f"{start}-{end}")
        else:
            # Single episode (e.g., "6")
            formatted_episodes.append(ep)

    # If more than 8 episodes, display only the first and last
    if len(formatted_episodes) > 8:
        first_episode = formatted_episodes[0]
        last_episode = formatted_episodes[-1]
        return f"{first_episode}-{last_episode}"  # Show the range

    # Otherwise, just return the formatted episodes
    return ', '.join(formatted_episodes)

@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        # Get form data
        date = request.form.get("date")
        episode = request.form.get("episode")
        title = request.form.get("title")
        language = request.form.get("language")
        duration = request.form.get("duration")
        location = request.form.get("watched_at")
        comment = request.form.get("comment") if request.form.get("leave_comment") == "yes" else "-"

        # Save data to Excel
        wb = openpyxl.load_workbook(FILE_PATH)
        ws = wb.active
        ws.append([date, episode, title, language, duration, location, comment])
        wb.save(FILE_PATH)

        return redirect(url_for("log"))

    return render_template("index1.html")

@app.route("/log")
def log():
    # Load data from Excel
    wb = openpyxl.load_workbook(FILE_PATH)
    ws = wb.active
    records = list(ws.iter_rows(values_only=True))[1:]  # Skip headers

    # Format episode range if applicable
    formatted_records = []
    for record in records:
        episode = record[1]
        # Use the format_episodes function to display formatted episode
        formatted_episode = format_episodes(episode)
        formatted_records.append(list(record)[:1] + [formatted_episode] + list(record)[2:])

    return render_template("log1.html", records=formatted_records)

@app.route("/edit/<int:index>", methods=["GET", "POST"])
def edit(index):
    wb = openpyxl.load_workbook(FILE_PATH)
    ws = wb.active
    record = list(ws.iter_rows(values_only=True))[1:][index]  # Get the record to edit

    if request.method == "POST":
        # Get form data
        date = request.form.get("date")
        episode = request.form.get("episode")
        title = request.form.get("title")
        language = request.form.get("language")
        duration = request.form.get("duration")
        location = request.form.get("watched_at")
        comment = request.form.get("comment") if request.form.get("leave_comment") == "yes" else "-"

        # Update the specific row in the worksheet
        row = index + 2  # Adjust the row index (because row 1 is the header)
        ws.cell(row=row, column=1, value=date)
        ws.cell(row=row, column=2, value=episode)
        ws.cell(row=row, column=3, value=title)
        ws.cell(row=row, column=4, value=language)
        ws.cell(row=row, column=5, value=duration)
        ws.cell(row=row, column=6, value=location)
        ws.cell(row=row, column=7, value=comment)

        wb.save(FILE_PATH)

        return redirect(url_for("log"))

    return render_template("index1.html", record=record)

@app.route("/delete/<int:index>")
def delete(index):
    # Load data from Excel
    wb = openpyxl.load_workbook(FILE_PATH)
    ws = wb.active
    ws.delete_rows(index + 2)  # Adjust index for header
    wb.save(FILE_PATH)

    return redirect(url_for("log"))

@app.route("/delete_all")
def delete_all():
    # Load data from Excel
    wb = openpyxl.load_workbook(FILE_PATH)
    ws = wb.active
    # Delete all rows except the header (index 1)
    ws.delete_rows(2, ws.max_row)
    wb.save(FILE_PATH)

    return redirect(url_for("log"))

if __name__ == "__main__":
    app.run(host='127.0.0.1', port=5001, debug=True)
